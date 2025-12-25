import pandas as pd
import openpyxl
import re
import os
import traceback
from datetime import datetime

class WHExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
            self.wb_formula = openpyxl.load_workbook(self.file_path, data_only=False)
            self._log(f"Loaded WH workbook: {file_path}")
        except Exception as e:
            self._log(f"Error loading WH workbook {file_path}: {e}")
            self.wb = None
            self.wb_formula = None
        self.route_options_cache = None

    def _log(self, msg):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [WH] {msg}\n"
        print(log_entry.strip())
        
        # Write to log file in working directory
        log_file = os.path.join(os.getcwd(), 'wh_cost_logs.txt')
        try:
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
        except Exception as e:
            print(f"Error writing to log file: {e}")

    def _find_header_info(self, ws):
        # Search first 5 rows for From and To
        for r in range(1, 6):
            row_values = [ws.cell(r, c).value for c in range(1, 20)]
            if 'From' in row_values and 'To' in row_values:
                info = {}
                for idx, val in enumerate(row_values):
                    if val:
                        info[str(val).strip()] = idx + 1
                return r, info
        return 2, {}

    def get_route_options(self):
        if self.route_options_cache:
            return self.route_options_cache
        
        if not self.wb or 'WAHL WH fee' not in self.wb.sheetnames:
            return {}
            
        ws = self.wb['WAHL WH fee']
        header_row, col_info = self._find_header_info(ws)
        from_col = col_info.get('From')
        to_col = col_info.get('To')
        
        if not from_col or not to_col:
            return {}
            
        options = {}
        route_to_node = {} # (from, to) -> NodeName
        node_counter = 0
        
        for r in range(header_row + 1, ws.max_row + 1):
            f_val = ws.cell(r, from_col).value
            t_val = ws.cell(r, to_col).value
            if not f_val or not t_val: continue
            
            frm = str(f_val).strip()
            to = str(t_val).strip()
            route_key = (frm, to)
            
            if route_key not in route_to_node:
                node_name = chr(ord('A') + node_counter)
                node_counter += 1
                route_to_node[route_key] = node_name
                options[node_name] = {
                    'locations': [f"{frm} -> {to}"],
                    'details': [], # Will store all alternate rows for this node
                    'sheet': 'WAHL WH fee'
                }
            
            curr_node = route_to_node[route_key]
            options[curr_node]['details'].append({
                'from': frm,
                'to': to,
                'excel_row': r,
                'own': str(ws.cell(r, col_info.get('Own', 1)).value or '')
            })
        
        self.route_options_cache = options
        return options

    INPUT_FIELDS = ['Import Truck times', 'Export Truck times', 'pallet', 'CBM', 'Month Qty', 'total invoice value (RMB)']

    def get_node_fields(self, node, location_str, current_inputs=None):
        options = self.get_route_options()
        if node not in options: return []
        
        ws = self.wb['WAHL WH fee']
        header_row, col_info = self._find_header_info(ws)
        details = options[node]['details']
        
        own_idx = col_info.get('Own', 1)
        invoice_idx = col_info.get('total invoice value (RMB)', 12)
        exclude = ['From', 'To', 'Method', 'TOTAL Cost(HKD)']

        # 1. Identify "Differentiators" - fixed for the node
        differentiators = []
        for c in range(own_idx, invoice_idx + 1):
            header = ws.cell(header_row, c).value
            header_str = str(header).strip() if header else ""
            if not header_str or header_str in exclude or header_str in self.INPUT_FIELDS:
                continue
            
            vals = set()
            for d in details:
                v = ws.cell(d['excel_row'], c).value
                if v is not None and str(v).strip().upper() != 'N/A' and str(v).strip() != '':
                    vals.add(str(v).strip())
            
            # Show field if it has ANY content (not just multiple values)
            if len(vals) > 0:
                all_opts = sorted(list(vals))
                print(f"[DEBUG] Node {node}: Field '{header_str}' has {len(all_opts)} options: {all_opts}")
                differentiators.append({
                    "name": header_str,
                    "col_idx": c,
                    "all_options": all_opts  # Store all possible values
                })

        print(f"[DEBUG] Node {node}: Found {len(details)} detail rows, {len(differentiators)} differentiator fields")
        
        fields = []
        # Differentiators ALWAYS show ALL their options (not filtered by other selections)
        for diff in differentiators:
            fields.append({
                "name": diff['name'],
                "display_name": diff['name'],
                "options": diff['all_options'],  # Always show all options
                "type": "select"
            })

        # 2. Get Input Fields - filtered based on ALL current selections
        matching_rows = details
        if current_inputs:
            print(f"[DEBUG] Filtering rows based on current inputs: {current_inputs}")
            for d_info in differentiators:
                val = current_inputs.get(d_info['name'])
                if val:
                    before_count = len(matching_rows)
                    matching_rows = [r for r in matching_rows if str(ws.cell(r['excel_row'], d_info['col_idx']).value or '').strip() == str(val).strip()]
                    print(f"[DEBUG]   Filter by {d_info['name']}='{val}': {before_count} -> {len(matching_rows)} rows")
        
        # If matching_rows is empty (conflict), we use all rows of node to avoid empty UI
        effective_rows = matching_rows if matching_rows else details
        print(f"[DEBUG] Using {len(effective_rows)} rows to determine available input fields")

        for header_str in self.INPUT_FIELDS:
            c = col_info.get(header_str)
            if not c: continue
            
            has_val = False
            for r in effective_rows:
                v = ws.cell(r['excel_row'], c).value
                if v is not None and str(v).strip().upper() != 'N/A':
                    has_val = True
                    break
            
            if has_val:
                val = ws.cell(effective_rows[0]['excel_row'], c).value
                fields.append({
                    "name": header_str,
                    "display_name": header_str,
                    "options": [str(val)] if val is not None else [],
                    "type": "input"
                })
            
        return fields

    def calculate(self, selections):
        results = []
        total_total_cost = 0
        
        self._log("=" * 60)
        self._log("WH CALCULATION START")
        self._log("=" * 60)
        
        ws = self.wb['WAHL WH fee']
        ws_formula = self.wb_formula['WAHL WH fee']
        header_row, col_info = self._find_header_info(ws)
        options = self.get_route_options()
        
        try:
            for idx, sel in enumerate(selections):
                node_id = sel.get('node')
                inputs = sel.get('inputs', {})
                if node_id not in options: continue
                
                # Attempt 1: Match on EVERYTHING provided in inputs
                self._log(f"--- Processing SECTION {idx+1} ---")
                details = options[node_id]['details']
                
                # First, log what we're trying to match
                self._log(f"  Attempting to match inputs: {inputs}")
                self._log(f"  Available rows for Node {node_id}: {len(details)} rows")
                
                matched_row = None
                for d in details:
                    r = d['excel_row']
                    match = True
                    for k, v in inputs.items():
                        c_idx = col_info.get(k)
                        if c_idx:
                            cell_val = str(ws.cell(r, c_idx).value or '').strip()
                            if cell_val != str(v).strip():
                                match = False
                                break
                    if match:
                        matched_row = r
                        self._log(f"  ✓ Exact match found at Excel Row {r}")
                        break
                
                # Attempt 2: Fallback - Match ignoring numeric INPUT_FIELDS
                if not matched_row:
                    self._log(f"  Strict match failed, trying fallback matching (ignoring numeric inputs)...")
                    self._log(f"  Numeric fields to ignore: {self.INPUT_FIELDS}")
                    
                    for d in details:
                        r = d['excel_row']
                        match = True
                        match_details = []
                        
                        for k, v in inputs.items():
                            if k in self.INPUT_FIELDS:
                                match_details.append(f"{k}=<skipped>")
                                continue  # Skip numeric inputs
                            
                            c_idx = col_info.get(k)
                            if c_idx:
                                cell_val = str(ws.cell(r, c_idx).value or '').strip()
                                expected_val = str(v).strip()
                                if cell_val == expected_val:
                                    match_details.append(f"{k}='{cell_val}'✓")
                                else:
                                    match_details.append(f"{k}: '{cell_val}'≠'{expected_val}'✗")
                                    match = False
                                    break
                        
                        self._log(f"    Row {r}: {', '.join(match_details)} -> {'MATCH' if match else 'NO MATCH'}")
                        
                        if match:
                            matched_row = r
                            self._log(f"  ✓ Fallback match found at Excel Row {r}")
                            break
                
                if not matched_row:
                    error_msg = f"无法找到匹配的Excel行。请检查您的选择组合。\n当前输入: {inputs}\n可用的行组合:"
                    for d in details:
                        r = d['excel_row']
                        row_info = []
                        for k in inputs.keys():
                            if k not in self.INPUT_FIELDS:
                                c_idx = col_info.get(k)
                                if c_idx:
                                    val = ws.cell(r, c_idx).value
                                    row_info.append(f"{k}='{val}'")
                        error_msg += f"\n  Row {r}: {', '.join(row_info)}"
                    
                    self._log(f"  [ERROR] {error_msg}")
                    raise ValueError(error_msg)

                self._log(f"  Selected Node: {node_id}")
                self._log(f"  Matched Excel Row: {matched_row}")
                
                cost_col = col_info.get('TOTAL Cost(HKD)')
                calc_cost, breakdown = self._recalculate_formula(ws, ws_formula, matched_row, cost_col, inputs, col_info, header_row)
                
                results.append({
                    "node": node_id,
                    "cost": calc_cost,
                    "breakdown": breakdown
                })
                total_total_cost += calc_cost
                self._log(f"  Section {idx+1} Cost: {calc_cost:.2f} HKD")
                
        except Exception as e:
            self._log(f"EXCEPTION: {str(e)}")
            self._log(traceback.format_exc())
            
        return {
            "node_results": results,
            "total_cost": total_total_cost
        }

    def _recalculate_formula(self, ws, ws_formula, row, col, user_inputs, col_info, header_row):
        formula = ws_formula.cell(row, col).value
        # If it's not a formula, just return the value
        if not formula or not isinstance(formula, str) or not formula.startswith('='):
            val = ws.cell(row, col).value
            return float(val) if val else 0.0, {"base": [], "variable": []}

        self._log(f"  Recalculating TOTAL Cost formula: {formula}")
        evaluated_val = self._evaluate_formula(ws, ws_formula, formula, row, user_inputs, col_info, header_row)
        self._log(f"  Final Calculated Result: {evaluated_val:.4f}")
        
        # Build breakdown for display
        # "把Own标题到total invoice value (RMB)标题之间有数据的列显示出来" -> These are inputs
        # "把每一个涉及到的单元格都看一遍 ... 把费用进行列表显示的内容也需要更新"
        # We need to show the cost items (Col 13 to 24) after recalculation
        breakdown = {"base": [], "variable": []}
        
        for c in range(13, 26): # Standard cost items start at Col 13
            header = ws.cell(header_row, c).value
            if not header: continue
            
            item_formula = ws_formula.cell(row, c).value
            if item_formula and isinstance(item_formula, str) and item_formula.startswith('='):
                val = self._evaluate_formula(ws, ws_formula, item_formula, row, user_inputs, col_info, header_row)
            else:
                val = ws.cell(row, c).value
            
            if val is not None and val != 0:
                # Row 3 contains the standard rates like '65000HKD/Month' etc.
                standard_rate = ws.cell(3, c).value
                breakdown["base"].append({
                    "name": str(header),
                    "row1": str(standard_rate) if standard_rate else (item_formula if item_formula else ""),
                    "row2": f"{val:.2f}" if isinstance(val, (int, float)) else str(val)
                })
        
        return evaluated_val, breakdown

    def _evaluate_formula(self, ws, ws_formula, formula, row, user_inputs, col_info, header_row):
        if not formula or not isinstance(formula, str): return 0.0
        if not formula.startswith('='):
            try: return float(formula)
            except: return 0.0
            
        expr = formula[1:].strip()
        
        # Handle SUM(Range)
        sum_matches = re.findall(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', expr)
        for start_col_str, start_row_str, end_col_str, end_row_str in sum_matches:
            start_col = openpyxl.utils.column_index_from_string(start_col_str)
            end_col = openpyxl.utils.column_index_from_string(end_col_str)
            s_row = int(start_row_str)
            e_row = int(end_row_str)
            
            sum_val = 0.0
            for r in range(s_row, e_row + 1):
                for c in range(start_col, end_col + 1):
                    cell_f = ws_formula.cell(r, c).value
                    if cell_f and isinstance(cell_f, str) and cell_f.startswith('='):
                        sum_val += self._evaluate_formula(ws, ws_formula, cell_f, r, user_inputs, col_info, header_row)
                    else:
                        header_name = self._get_header_name(ws, header_row, c)
                        if header_name and header_name in user_inputs:
                            try: sum_val += float(user_inputs[header_name])
                            except: pass
                        else:
                            val = ws.cell(r, c).value
                            try: sum_val += float(val) if val else 0.0
                            except: pass
            
            expr = expr.replace(f'SUM({start_col_str}{start_row_str}:{end_col_str}{end_row_str})', str(sum_val))

        # Handle Cell references
        cell_refs = re.findall(r'([A-Z]+)(\d+)', expr)
        # Sort by length descending to avoid replacing A10 before A1
        cell_refs.sort(key=lambda x: len(x[0]+x[1]), reverse=True)
        
        for col_str, row_str in cell_refs:
            c_idx = openpyxl.utils.column_index_from_string(col_str)
            r_idx = int(row_str)
            
            header_name = self._get_header_name(ws, header_row, c_idx)
            val = None
            
            if header_name and header_name in user_inputs and r_idx == row:
                try: 
                    val = float(user_inputs[header_name])
                    self._log(f"      Replacing {col_str}{row_str} with User Input '{header_name}': {val}")
                except: 
                    val = 0.0
            else:
                cell_f = ws_formula.cell(r_idx, c_idx).value
                if cell_f and isinstance(cell_f, str) and cell_f.startswith('='):
                    val = self._evaluate_formula(ws, ws_formula, cell_f, r_idx, user_inputs, col_info, header_row)
                    self._log(f"      Formula at {col_str}{row_str} evaluated to: {val}")
                else:
                    v = ws.cell(r_idx, c_idx).value
                    try: 
                        val = float(v) if v else 0.0
                        self._log(f"      Value at {col_str}{row_str} is: {val}")
                    except: 
                        val = 0.0
            
            # Match whole word to avoid replacing A10 when we meant A1
            pattern = r'\b' + col_str + row_str + r'\b'
            if re.search(pattern, expr):
                expr = re.sub(pattern, str(val), expr)
                self._log(f"      [Formula Step] {col_str}{row_str} -> {val}")

        try:
            # Basic sanitization
            safe_expr = expr.replace('^', '**')
            # Only allow mathematical ops
            if all(c in '0123456789.+-*/() *' for c in safe_expr):
                result = eval(safe_expr)
                return float(result)
            else:
                self._log(f"Warning: Unsafe expression detected: {expr}")
                return 0.0
        except Exception as e:
            self._log(f"Error evaluating {expr}: {e}")
            return 0.0

    def _get_header_name(self, ws, header_row, col):
        return str(ws.cell(header_row, col).value).strip() if ws.cell(header_row, col).value else None
